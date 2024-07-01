import pandas as pd
import numpy as np
import argparse
import os
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def graph(input_dir, filename, threshold):
    if filename.startswith('~$'):
        print(f"Skipping temporary file: {filename}")
        return []

    input_file_path = os.path.join(input_dir, filename)
    df = pd.read_excel(input_file_path)

    war_mixed_list = []
    war_proc_list = []

    for index, row in df.iterrows():
        if row['File index'] != 'Average':  # Skip appending the last row
            war_mixed_list.append(row['Mixed W.A.R. (%)'])
            war_proc_list.append(row['Processed W.A.R. (%)'])

    mixed_avg = round(np.average(war_mixed_list))
    proc_avg = round(np.average(war_proc_list))

    results = []
    expand_res = []

    for i in range(90, -1, -10):
        keys_over_i = [war_proc for war_mixed, war_proc in zip(war_mixed_list, war_proc_list) if (war_mixed >= i and war_mixed < (i + (11 if i == 90 else 10)))]
        if threshold != -1:
            count_ge_threshold = np.sum(np.array(keys_over_i) >= threshold)
        else:
            count_ge_threshold = np.sum(np.array(keys_over_i) < 70)

        if count_ge_threshold == 0 or len(keys_over_i) == 0:
            percentage_ge_threshold = 0
        else:
            percentage_ge_threshold = (count_ge_threshold / len(keys_over_i)) * 100

        results.append(percentage_ge_threshold)
        expand_res.append((count_ge_threshold, len(keys_over_i)))

    return results, expand_res, mixed_avg, proc_avg

def read_excel_without_last_row(file_path):
    df = pd.read_excel(file_path)
    return df.iloc[:-1]  # Exclude the last row

def highlight_columns(writer, sheet_name):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for col in ['D', 'G']:  # Highlight columns D (Average Improvement) and G (Figure of Merit)
        for cell in worksheet[col]:
            cell.fill = highlight_fill

def highlight_row(writer, sheet_name, row_index):
    workbook = writer.book
    worksheet = writer.sheets[sheet_name]
    highlight_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

    for row in worksheet.iter_rows(min_row=row_index, max_row=row_index): 
        for cell in row:
            cell.fill = highlight_fill

def main():
    parser = argparse.ArgumentParser(description='Process Excel files and generate a stacked bar graph.')
    parser.add_argument('input_dir', type=str, help='Directory where input Excel files are located')
    args = parser.parse_args()

    input_dir = args.input_dir
    excel_files = [file for file in os.listdir(input_dir) if file.endswith('.xlsx') or file.endswith('.xls')]

    results_table = []

    for file in excel_files:
        df = read_excel_without_last_row(os.path.join(input_dir, file))

        thresholds = [90, 80, 70, -1]
        colors = ['r', 'g', 'b', 'k']

        fig, ax = plt.subplots(figsize=(10, 6))

        width = 0.2  # width of the bars
        x = np.arange(10)  # the label locations
        x_labels = ['90%', '80%', '70%', '60%', '50%', '40%', '30%', '20%', '10%', '0%']

        sum_of_red = 0
        denom_of_90 = 0

        file_results = {
            'filename': file,
            'High Processed Tail': 0,
            'High Mixed Tail': 0,
            'Figure of Merit': 0,
            'Average Mixed WAR': 0,
            'Average Processed WAR': 0,
            'Average Improvement': 0
        }

        for i, threshold in enumerate(thresholds):
            results, expand_res, mixed_avg, proc_avg = graph(input_dir, file, threshold)
            denom_of_90 = expand_res[0][1]
            bars = ax.bar(x + i * width, results, width, label=(f'WAR >= {threshold}%' if threshold != -1 else "WAR < 70%"), color=colors[i])
            for bar, result in zip(bars, expand_res):
                count_ge_i, total_keys_i = result
                ax.text(bar.get_x() + bar.get_width() / 2, bar.get_height() + 1, f"{count_ge_i}/{total_keys_i}", ha='center', va='bottom')
                if threshold == 90:
                    sum_of_red += count_ge_i

        FOM = round(((sum_of_red / denom_of_90) - 1) * 100)
        file_results['High Processed Tail'] = sum_of_red
        file_results['High Mixed Tail'] = denom_of_90
        file_results['Figure of Merit'] = FOM
        file_results['Average Mixed WAR'] = mixed_avg
        file_results['Average Processed WAR'] = proc_avg
        file_results['Average Improvement'] = proc_avg - mixed_avg

        results_table.append(file_results)

        ax.set_ylabel('Processed W.A.R. %')
        ax.set_xlabel('W.A.R. on Original')
        ax.set_title(f'Bar Graph of Processed WAR% >= Thresholds vs. Original WAR% for {file}')
        ax.set_xticks(x + width)
        ax.set_xticklabels(x_labels)
        ax.legend()
        ax.set_ylim(0, 120)
        plt.tight_layout()
        plt.show()

    # Convert results to a DataFrame and sort by filename
    results_df = pd.DataFrame(results_table)
    results_df.sort_values(by='filename', ascending=True, inplace=True)

    # Find the row with the closest Average Mixed WAR to 70%
    closest_row_index = (results_df['Average Mixed WAR'] - 70).abs().idxmin()
    # print(closest_row_index)
    # Reorder columns
    results_df = results_df[['filename', 'Average Mixed WAR', 'Average Processed WAR', 'Average Improvement', 'High Mixed Tail', 'High Processed Tail', 'Figure of Merit']]

    # Save the DataFrame to an Excel file
    output_excel = os.path.join(input_dir, 'results_summary.xlsx')
    with pd.ExcelWriter(output_excel, engine='openpyxl') as writer:
        results_df.to_excel(writer, index=False, sheet_name='Summary')
        highlight_columns(writer, 'Summary')
        highlight_row(writer, 'Summary', closest_row_index)

    print(f"Results have been saved to {output_excel}")

if __name__ == "__main__":
    main()
